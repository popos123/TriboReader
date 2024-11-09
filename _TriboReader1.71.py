import re
import os
import time
import ctypes
import pandas as pd
import msvcrt  # Importuj bibliotekę do obsługi klawiszy
from scipy.signal import savgol_filter, medfilt
from openpyxl import Workbook
from openpyxl.chart import Reference, Series, ScatterChart, LineChart
import configparser

# Enable ANSI escape codes in terminal
kernel32 = ctypes.windll.kernel32
kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)

# Program wczytuje wszystkie pliki z tribometru i generuje wykresy w pliku .xlsx
# Nazwa pliku + .txt lub .csv to nazwa wykresu
# Jeśli w nazwie pliku jest nawias to tekst w nim zawarty będzie opisem osi, dla title_from_text = 0
# Dla Rtec w nazwie musi być zawarta prędkość liniowa (np. 0.1m-s), a dla T11 dodatkowo obciążenie (np. 10N)

def load_config(file_name):
    config = configparser.ConfigParser()
    config.read(file_name)
    settings = config["Settings"]
    return {
        "offset_raw": int(settings.get("offset_raw", 1)),
        "title_from_text": int(settings.get("title_from_text", 0)),
        "erase_peak": int(settings.get("erase_peak", 0)),
        "invert_peak": int(settings.get("invert_peak", 1)),
        "default_window_length_u": int(settings.get("default_window_length_u", 7)),
        "default_window_length_pd": int(settings.get("default_window_length_pd", 5)),
        "min_sample": int(settings.get("min_sample", 100)),
        "max_sample": int(settings.get("max_sample", 500))
    }

def ask_user_for_variables():
    pd_set = 0
    offset_raw, erase_peak, invert_peak = 1, 0, 1
    title_from_text = 0
    default_window_length_u, default_window_length_pd = 7, 5
    min_sample, max_sample = 100, 500
    while True:
        offset_raw_input = input(f"Czy offsetować dane RAW zużycia liniowego? [1 - tak, 0 - nie] ({offset_raw}): ").strip() or '1'
        title_from_text_input = input(f"Czy nazwać serię danych tekstem z nawiasu nazwy pliku? [1 - tak, 0 - nie] ({title_from_text}): ").strip() or '1'
        pd_set_input = input(f"Czy dane zużycia liniowego z peakiem na początku: ucinać (1), odwracać (2), nic nie robić (3)? ({pd_set}): ").strip() or '2'
        min_sample_input = input(f"Minimalna ilość danych dla filtru Savitzky-Golay ({min_sample}): ").strip() or str(min_sample)
        max_sample_input = input(f"Maksymalna ilość danych dla filtru Savitzky-Golay ({max_sample}): ").strip() or str(max_sample)
        window_length_u_input = input(f"Długość okna filtru Savitzky-Golay dla µ ({default_window_length_u}): ").strip() or str(default_window_length_u)
        window_length_pd_input = input(f"Długość okna filtru Savitzky-Golay dla pd ({default_window_length_pd}): ").strip() or str(default_window_length_pd)
        try:
            offset_raw = int(offset_raw_input)
            title_from_text = 1 - int(title_from_text_input) # inwersja
            pd_set = int(pd_set_input)
            min_sample = int(min_sample_input)
            max_sample = int(max_sample_input)
            window_length_u = int(window_length_u_input)
            window_length_pd = int(window_length_pd_input)
            if window_length_u > int(max_sample_input) or window_length_pd > int(max_sample_input):
                print(f"\033[38;5;214m Liczba większa niż {int(max_sample_input)} \033[0m")
                continue # Powrót na początek pętli
            default_window_length_u = window_length_u
            default_window_length_pd = window_length_pd
            break
        except ValueError:
            print(f"\033[91m Błąd: Wprowadź poprawną liczbę mniejszą niż {int(max_sample_input)} \033[0m")
    if pd_set == 1: erase_peak, invert_peak = 1, 0
    if pd_set == 2: erase_peak, invert_peak = 0, 1
    if pd_set == 3: erase_peak, invert_peak = 0, 0
    return min_sample, max_sample, default_window_length_u, default_window_length_pd, title_from_text, offset_raw, erase_peak, invert_peak

def Savitzky(averaged_data, default_window_length_u, default_window_length_pd):
    # Apply median filter to 'Penetration Depth [µm]'and 'µ' column within oscillation ranges
    # Calculate the maximum allowable window length
    max_window_length = (len(averaged_data) // 2) * 2 - 1  # Ensures it's odd
    # Use the minimum of the default window length and the maximum allowable window length
    window_length_u = min(default_window_length_u, max_window_length)
    window_length_pd = min(default_window_length_pd, max_window_length)
    # Ensure the window length is odd
    if window_length_u % 2 == 0:
        window_length_u += 1
    if window_length_pd % 2 == 0:
        window_length_pd += 1

    first_value_pd = averaged_data.at[0, 'Penetration Depth [µm]']
    second_value_pd = averaged_data.at[1, 'Penetration Depth [µm]']
    first_value_u = averaged_data.at[0, 'µ']
    
    #averaged_data['Penetration Depth [µm]'] = medfilt(df['Penetration Depth [µm]'], window_length_pd)
    #averaged_data['µ'] = medfilt(df['µ'], window_length_u)

    averaged_data['Penetration Depth [µm]'] = savgol_filter(averaged_data['Penetration Depth [µm]'], window_length_pd, polyorder=2)
    averaged_data['µ'] = savgol_filter(averaged_data['µ'], window_length_u, polyorder=2)
    
    averaged_data.at[0, 'Penetration Depth [µm]'] = first_value_pd
    averaged_data.at[1, 'Penetration Depth [µm]'] = second_value_pd
    averaged_data.at[0, 'µ'] = first_value_u

    return averaged_data

# Funkcja konwertuje przebieg Penetration Depth [µm] z pseudo sinusoidalnego / prostokątnego na liniowy dodatni
def linear_mode_u_preprocessing(df):
    linear_positions = df['Linear Position [mm]']
    sign_changes_linear = (linear_positions.shift() * linear_positions < 0).cumsum()
    # Initialize sign_changes_µ and other necessary variables
    sign_changes_µ = pd.Series(index=df.index)
    last_sign_change_index = None
    max_forward_limit = 10
    for idx in df.index:
        if sign_changes_linear[idx] != last_sign_change_index:
            last_sign_change_index = sign_changes_linear[idx]
            sign_change_count = 0
        if (linear_positions.shift()[idx] * df['µ'][idx] < 0) or (sign_change_count < max_forward_limit):
            sign_change_count += 1
        else:
            sign_change_count = max_forward_limit 
        sign_changes_µ[idx] = sign_change_count
    # Group by 'sign_changes_linear'
    df['sign_changes_linear'] = sign_changes_linear
    # Initialize the resulting column
    df['shifted_µ'] = 0.0  # Set float type for the shifted_µ column
    # Group by 'sign_changes_linear' and calculate the shift
    groups = df.groupby('sign_changes_linear')
    for name, group in groups:
        if name == 0:
            df.loc[group.index, 'shifted_µ'] = group['µ']
        else:
            prev_value = group['µ'].iloc[0]
            median_sign = -1 if prev_value < 0 else 1
            df.loc[group.index, 'shifted_µ'] = group['µ'] * median_sign
    # Remove the 'µ' column and rename 'shifted_µ' to 'µ'
    df.drop(columns=['µ'], inplace=True)
    df.rename(columns={'shifted_µ': 'µ'}, inplace=True)
    # # Calculate the required number of samples for filtering
    # # Determine the default window length
    # default_window_length = 7
    # # Calculate the maximum allowable window length
    # max_window_length = (len(df) // 2) * 2 - 1  # Ensures it's odd
    # # Use the minimum of the default window length and the maximum allowable window length
    # window_length = min(default_window_length, max_window_length)
    # # Ensure the window length is odd
    # if window_length % 2 == 0:
    #     window_length += 1
    sampling_percent = 0.12  # 12%
    # Identify the indices where sign_changes_linear changes
    change_indices = pd.Series(sign_changes_linear).diff().ne(0).cumsum() - 1
    change_indices = change_indices[change_indices != -1].index.tolist()
    # Calculate the required number of samples for filtering
    num_samples_before = int((len(df) / len(change_indices)) * sampling_percent)
    num_samples_after = int((len(df) / len(change_indices)) * sampling_percent * 1.5)
    # Apply approximation within the specified ranges based on sign_changes_linear changes
    for idx in change_indices:
        start_idx = max(0, idx - num_samples_before)
        end_idx = min(len(df) - 1, idx + num_samples_after)
        filter_range = range(start_idx, end_idx + 1)
        # Ensure there are at least two elements to approximate
        if len(filter_range) >= 2:
            if len(filter_range) > 4:
                first_two = df.loc[filter_range, 'µ'].iloc[:2].tolist()
                last_two = df.loc[filter_range, 'µ'].iloc[-2:].tolist()
                approx_values = first_two + last_two
                df.loc[filter_range, 'µ'] = approx_values * (len(filter_range) // 4) + approx_values[:len(filter_range) % 4]
            else:
                # Use all available values if there are only two elements
                df.loc[filter_range, 'µ'] = df.loc[filter_range, 'µ'].iloc[:2].tolist() * (len(filter_range) // 2) + df.loc[filter_range, 'µ'].iloc[:len(filter_range) % 2].tolist()
    #     else:
    #         # Wygładzenie danych 'µ'
    #         df['µ'] = savgol_filter(df['µ'], window_length, polyorder=2)
    # # Apply Savitzky-Golay filter to 'µ' column within oscillation ranges
    # filtered_data = savgol_filter(df['µ'], window_length, polyorder=2)
    # min_filtered_value = filtered_data.min()
    # if min_filtered_value <= 0:
    #     df['µ'] = abs(filtered_data)
    # else:
    #     # Shift the filtered data upwards so that the minimum value is zero or greater
    #     shifted_data = filtered_data + abs(min_filtered_value)
    #     df['µ'] = shifted_data
    return df

# Wczytanie dla tribometru T11 prędkości liniowej i obciążenia z nazwy pliku z danymi, obliczenie µ i Distance [m]
def T11_calculations(df, file_name):
    s = 0.1 # Default
    F = 10 # Default
    # Dodanie kolumn µ i Distance [m]
    df["µ"] = 0.0  # Placeholder na współczynnik tarcia, do obliczenia poniżej
    df["Distance [m]"] = 0.0  # Placeholder na drogę, do obliczenia poniżej

    # Obliczenie prędkości liniowej i drogi (Distance [m])
    try:
        # Szukanie wartości s z nazwy pliku
        match = re.search(r'(?<!\d)(\d+(?:[.,]\d+)?)(?=\s?(?:m-s|ms|m\s?-\s?s))', file_name) # Obsługa liczb z kropką i przecinkiem
        if match:
            s = float(match.group(1).replace(',', '.'))
            print(f"[T11] Odczytana prędkość liniowa: {s} m/s")
        else:
            # TODO Podanie ręcznie jak nie odczyta z nazwy pliku
            raise ValueError("Brak odpowiedniej wartości m/s w nazwie pliku, np. 0.1m-s [zastosowano domyślnie 0.1m/s]")
        # Obliczanie całkowitej przebytej drogi (m) przy stałej prędkości z każdej chwili czasowej
        df['Distance [m]'] = s * df['Time [s]']
    except Exception as e:
        raise ValueError(f"Problem z obliczeniem drogi: {e}")

    # Obliczenie współczynnika tarcia µ
    try:
        # Szukanie wartości F z nazwy pliku
        match = re.search(r'(?<!\d)(\d+(?:[.,]\d+)?)(?=\s?[Nn])', file_name) # Obsługa liczb z kropką i przecinkiem
        if match:
            F = float(match.group(1).replace(',', '.'))
            print(f"[T11] Odczytana wartość obciążenia: {F} N")
        else:
            # TODO Podanie ręcznie jak nie odczyta z nazwy pliku
            raise ValueError("Brak odpowiedniej wartości F w nazwie pliku, np. 10N [zastosowano domyślnie 10N]")
        # Obliczanie µ = N/F, gdzie N to "Friction force [N]", F odczytane z pliku
        df["µ"] = df["Friction force [N]"].apply(lambda N: N / F if N != 0 else 0)
    except Exception as e:
        raise ValueError(f"Problem z obliczeniem współczynnika tarcia µ: {e}")
    return df

# Wczytanie dla tribometru Rtec prędkości liniowej z nazwy pliku z danymi, konwersja Penetration Depth [µm] i Distance [m]
def Rtec_calculations(df, file_name):
    s = 0.1 # Default
    # Dodanie kolumn Penetration Depth [µm] i Distance [m]
    df["Penetration Depth [µm]"] = 0.0  # Placeholder na zużycie liniowe, do obliczenia poniżej
    df["Distance [m]"] = 0.0  # Placeholder na drogę, do obliczenia poniżej

    # Obliczenie zuźycia liniowego i drogi (Distance [m])
    try:
        # Szukanie wartości s z nazwy pliku
        match = re.search(r'(?<!\d)(\d+(?:[.,]\d+)?)(?=\s?(?:m-s|ms|m\s?-\s?s))', file_name)# Obsługa liczb z kropką i przecinkiem
        if match:
            s = float(match.group(1).replace(',', '.'))
            print(f"[Rtec] Odczytana prędkość liniowa: {s} m/s")
        else:
            # TODO Podanie ręcznie jak nie odczyta z nazwy pliku
            raise ValueError("Brak odpowiedniej wartości m/s w nazwie pliku, np. 0.1m-s [zastosowano domyślnie 0.1m/s]")
        # Obliczanie całkowitej przebytej drogi (m) przy stałej prędkości z każdej chwili czasowej
        df['Distance [m]'] = s * df[' Timestamp']
    except Exception as e:
        raise ValueError(f"Problem z obliczeniem drogi: {e}")

    # Obliczenie Penetration Depth [µm]
    try:        
        # Obliczanie Penetration Depth [µm] z XYZ.Z Position (mm) poprzez pomnożenie przez 1000 aby z mm zrobić µm
        df["Penetration Depth [µm]"] = df["XYZ.Z Position (mm)"] * 1000.0
    except Exception as e:
        raise ValueError(f"Problem z obliczeniem zużycia liniowego: {e}")
    return df

# Główna funkcja wczytująca pliki i dane do DataFrame (df)
# 1. Otwiera plik w trybie odczytu, z kodowaniem UTF-8 i możliwością zastępowania błędów, 
# 2. Wczytuje wszystkie linie z pliku do listy, usuwa znak BOM, 
# 3. Definiowanie typów tribometrów i rodzajów trybu pracy, 
# 4. Wyszukiwanie nagłówka i przypisanie tribometru i rodzaju po nagłówku, 
# 5. Walidacja danych, usunięcie nieprawidłowych, utworzenie DataFrame,
# 6. Podział danych na tribometry i obróbka według każdego rodzaju tribometru i typu
def read_and_process_file(file_path):
    try:
        # Otwórz pliki
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            lines = file.readlines()

        # Usuń BOM z pierwszej linii, jeśli jest obecny (tylko w UTF-8)
        if lines[0].startswith('\ufeff'):
            lines[0] = lines[0][1:]
        
        # Zdefiniuj możliwe typy tribometrów i trybów
        # tribometer_types = ["Nano", "T11", "TRB3", "Rtec"]
        tribometer_types = ["\033[38;5;214mNano\033[0m", "\033[38;5;214mT11\033[0m", "\033[38;5;214mTRB3\033[0m", "\033[38;5;214mRtec\033[0m"] # Nano, T11, TRB3, Rtec
        modes = ["\033[38;5;208mLinear\033[0m", "\033[38;5;208mRotary\033[0m"] # Linear, Rotary

        # Znajdź linię z nagłówkiem i ustal początek danych (ich indeks)
        start_index = None
        tribometer_type = ""
        mode = ""
        
        # Początek danych, typ tribometru i tryb
        for i, line in enumerate(lines):
            # Tryb
            if "Linear mode" in line:
                mode = modes[0]  # "Linear"
            elif "Single-way mode" in line:
                mode = modes[1]  # "Rotary"
            # Typ
            if "Nano Tribometer" in line:
                tribometer_type = tribometer_types[0]  # "Nano"
            elif "TRB3" in line:
                tribometer_type = tribometer_types[2]  # "TRB3"
            # Tryb i typ i index
            # Nano Tribometer
            if "Time [s]\tDistance [m]\tlaps\tSequence ID\tCycle ID\tMax linear speed [m/s]\tNominal Load [mN]\tµ\tAngle [°]\tNormal force [mN]\tFriction force [mN]\tPenetration depth [µm]" in line:
                mode = modes[1]  # "Rotary" - bo jest Angle [°] w headerze
                tribometer_type = tribometer_types[0]  # "Nano" - bo jest jednostka "mN" w headerze
                start_index = i # i czyli zaczyna od headera
            elif "Time [s]\tDistance [m]\tlaps\tSequence ID\tCycle ID\tMax linear speed [m/s]\tNominal Load [mN]\tµ\tLinear Position [mm]\tNormal force [mN]\tFriction force [mN]\tPenetration depth [µm]" in line:
                mode = modes[0]  # "Linear" - bo jest Linear Position [mm] w headerze
                tribometer_type = tribometer_types[0]  # "Nano" - bo jest jednostka "mN" w headerze
                start_index = i # i czyli zaczyna od headera
            # TRB3
            elif "Time [s]\tDistance [m]\tLaps\tSequence ID\tCycle ID\tMax Linear Speed [m/s]\tNominal Load [N]\tµ\tAngle [°]\tFriction Force [N]\tTemperature [°C]\tHumidity [%]\tPenetration Depth [µm]" in line:
                mode = modes[1]  # "Rotary" - bo jest Angle [°] w headerze
                tribometer_type = tribometer_types[2]  # "TRB3" - bo jest jednostka "N" oraz Temperature [°C] i Humidity [%] w headerze
                start_index = i # i czyli zaczyna od headera
            elif "Time [s]\tDistance [m]\tLaps\tSequence ID\tCycle ID\tMax Linear Speed [m/s]\tNominal Load [N]\tµ\tLinear Position [mm]\tFriction Force [N]\tTemperature [°C]\tHumidity [%]\tPenetration Depth [µm]" in line:
                mode = modes[0]  # "Linear" - bo jest Linear Position [mm] w headerze
                tribometer_type = tribometer_types[2]  # "TRB3" - bo jest jednostka "N" oraz Temperature [°C] i Humidity [%] w headerze
                start_index = i # i czyli zaczyna od headera
            # T11
            elif "Time [s];Friction force [N];Displacement [um];Temperature2 [C];Temperature1 [C];Rotational speed [rpm];Number of revolutions" in line:
                mode = modes[1]  # "Rotary" - bo jest Number of revolutions w headerze
                tribometer_type = tribometer_types[1]  # "T11" - bo separator ";" oraz Temperature1 [C] i Temperature2 [C]
                start_index = i # i czyli zaczyna od headera
            # Rtec
            elif "Step, Timestamp, RecipeStep, DAQ.Fz (N),DAQ.Fx (N),DAQ.COF (),Rotary.Velocity (rpm),XYZ.Z Depth (mm),XYZ.Z Position (mm),Rotary.Angle (deg)," in line:
                mode = modes[1]  # "Rotary" - bo jest Rotary.Angle (deg) w headerze
                tribometer_type = tribometer_types[3]  # "Rtec" - bo jest separator "," oraz inne nagłówki i jednostki w "()"
                start_index = i # i czyli zaczyna od headera
        
        if start_index is None:
            raise ValueError(f"\033[91m Nie znaleziono odpowiedniej linii rozpoczynającej dane w pliku: {file_path} \033[0m")
        
        # Wczytaj nagłówek i dane
        # HEADER
        # .rstrip() - Usuń znaki końca linii i separator na końcu jak są
        # Dla T11 ";" jako separator
        if tribometer_type == tribometer_types[1]:  # "T11"
            header = lines[start_index].strip().rstrip(';\r\n').split(';')
        # Dla TRB3 i Nano TRB to TAB (\t) jako separator
        elif tribometer_type == tribometer_types[2] or tribometer_type == tribometer_types[0]:  # "TRB3 i Nano"
            header = lines[start_index].strip().rstrip('\t\r\n').split('\t')
        # Dla Rtec "," jako separator
        elif tribometer_type == tribometer_types[3]:  # "Rtec"
            header = lines[start_index].strip().rstrip(',\r\n').split(',')
        # W innym przypadku lub błędu
        else:
            raise ValueError(f"\033[91m Nie znaleziono separatora nagłówków \033[0m")

        # DANE
        # Dla T11 i Rtec dodatkowy poniżej fix na usunięcie znaku separatora na końcu linii i usunięcie linii z wartościami zerowymi w kolumnie "Friction force [N]"
        data = []
        valid_rows = 0
        invalid_rows = 0
        for line in lines[start_index + 1:]:
            # Dla T11 ";" jako separator
            if tribometer_type == tribometer_types[1]:  # "T11"
                line = line.strip().rstrip('\r\n;') # Usuń znaki końca linii i separator na końcu jak są
                row = line.strip().split(';')
            # Dla Rtec "," jako separator
            elif tribometer_type == tribometer_types[3]:  # "Rtec"
                line = line.strip().rstrip('\r\n,') # Usuń znaki końca linii i separator na końcu jak są
                row = line.strip().split(',')
            # Dla TRB3 i Nano TRB to TAB (\t) jako separator
            elif tribometer_type == tribometer_types[2] or tribometer_type == tribometer_types[0]:  # "TRB3 i Nano"
                line = line.strip().rstrip('\r\n\t') # Usuń znaki końca linii i separator na końcu jak są
                row = line.strip().split('\t')
            # W innym przypadku lub błędu
            else:
                raise ValueError(f"\033[91m Nie znaleziono separatora danych \033[0m")
            
            # Walidacja, czy wiersz zawiera odpowiednią liczbę kolumn, jak nie to pomiń
            if len(row) != len(header):
                invalid_rows += 1
                continue

            # Walidacja, czy wszystkie elementy można skonwertować na float (zamiana ewentualnie , na .)
            try:
                #row = [float(value.replace(',', '.')) for value in row]
                #data.append(row)
                #liczba_poprawnych_wierszy += 1
                for i in range(len(row)):
                    if ',' in row[i]: # jeśli przecinek
                        row[i] = row[i].replace(',', '.') # konwertuj na kropkę
                    if not re.match(r'^-?\d+(?:\.\d+)?$', row[i]): # jeśli jest tekst to pomijaj linię
                        invalid_rows += 1
                        break
                    else:
                        row[i] = float(row[i]) # jeśli wszystko w porządku, konwertuj na float
                else:
                    data.append(row) # przypisz do nowej macierzy dane wyselekcjonowane / poprawne
                    valid_rows += 1

            except ValueError as e:
                #print(f"\033[91m Błąd konwersji wartości na float w wierszu: {row} \033[0m")
                invalid_rows += 1
                continue
        
        if invalid_rows > 0:
            print(f"[{tribometer_type}] \033[91mLiczba niepoprawnych wierszy: {invalid_rows} z {valid_rows + invalid_rows}\033[0m")

        # Usuń wiersze z brakującymi lub nieprawidłowymi danymi
        data = [row for row in data if all(value != '' and value is not None for value in row)]

        # Sprawdź, czy są jakieś poprawne dane do przetworzenia
        if not data:
            raise ValueError(f"\033[91m Brak poprawnych danych w pliku {file_path} \033[0m")

        # Stwórz DataFrame
        df = pd.DataFrame(data, columns=header)

        # Sprawdź, czy plik zawiera kolumnę "Linear Position [mm]" jak tak to ruch Posuwisto-Zwrotny
        if mode == modes[0]:  # "Linear"
            if tribometer_type == tribometer_types[2] or tribometer_type == tribometer_types[0]:  # "TRB3 i Nano"
                df = linear_mode_u_preprocessing(df)
                
                # Wybierz interesujące kolumny
                df = df[['Distance [m]', 'µ', 'Penetration Depth [µm]']]
            else:
                raise ValueError(f"\033[91m Brak funkcji programu typu: {mode}, dla tego tribometru: {tribometer_type} \033[0m")
        # Jak nie to ruch Obrotowy
        elif mode == modes[1]:  # "Rotary"
            # Jeśli jest to Nano TRB
            if tribometer_type == tribometer_types[0]:  # "Nano"
                df['Penetration Depth [µm]'] = 0 # Set 0 to all rows in this column
                # Obliczenie średniej ucinanej dla wartości bez pików
                mean = df['µ'].mean()
                max_mean_range = 2.0
                if mean >= 0:
                    max_cut_out = mean + max_mean_range # maksymalna wartość średniej ucinanej
                else:
                    max_cut_out = abs(mean) + max_mean_range # maksymalna wartość średniej ucinanej
                mean_value = df.loc[(df['µ'] <= max_cut_out) & (df['µ'] >= -max_cut_out), 'µ'].mean()
                # Zastąpienie pików średnią
                df['µ'] = df['µ'].apply(lambda x: mean_value if x > max_cut_out or x < -max_cut_out else x)

                # Wybierz interesujące kolumny
                df = df[['Distance [m]', 'µ', 'Penetration Depth [µm]']]
            # Jeśli to T11
            elif tribometer_type == tribometer_types[1]:  # "T11"
                # Usunięcie błędnych odczytów
                # Zachowaj pierwszą wartość i usuń pozostałe wiersze gdzie Friction force = 0
                first_row = df.iloc[0]
                df = df.iloc[1:]  # Usuń pierwszy wiersz
                df = df[df['Friction force [N]'] != 0]  # Usuń wiersze z zerowymi wartościami
                df = pd.concat([pd.DataFrame([first_row]), df])  # Dodaj pierwszy wiersz z powrotem

                # Znajdź najniższą wartość w kolumnie "Friction force [N]"
                min_value = df['Friction force [N]'].min()
                # Jeśli najniższa wartość jest mniejsza od zera, zrób offset
                if min_value < 0:
                    df['Friction force [N]'] = df['Friction force [N]'] - min_value

                # Zmień nazwy kolumny
                df = df.rename(columns={'Displacement [um]': 'Penetration Depth [µm]'})
                # Dodanie kolumn "Distance [m]" oraz "µ" i obliczenie tych danych dla Tribometru T11
                df = T11_calculations(df, os.path.basename(file_path))
                
                # Wybierz interesujące kolumny
                df = df[['Distance [m]', 'µ', 'Penetration Depth [µm]']]
            # Jeśli to TRB3
            elif tribometer_type == tribometer_types[2]:  # "TRB3"
                # Znajdź najniższą wartość w kolumnie "µ"
                min_value = df['µ'].min()
                # Jeśli najniższa wartość jest mniejsza od zera, zrób offset
                if min_value < 0:
                    df['µ'] = df['µ'] - min_value

                # Wybierz interesujące kolumny
                df = df[['Distance [m]', 'µ', 'Penetration Depth [µm]']]
            # Jeśli to Rtec
            elif tribometer_type == tribometer_types[3]:  # "Rtec"
                # Usunięcie błędnych odczytów
                # Zachowaj pierwszą wartość i usuń pozostałe wiersze gdzie DAQ.COF () = 0
                first_row = df.iloc[0]
                df = df.iloc[1:]  # Usuń pierwszy wiersz
                df = df[df['DAQ.COF ()'] != 0]  # Usuń wiersze z zerowymi wartościami
                df = pd.concat([pd.DataFrame([first_row]), df])  # Dodaj pierwszy wiersz z powrotem

                # # Obliczenie średniej ucinanej dla wartości bez pików - wymaga bo Rtec ma dużo pików
                # mean = df['DAQ.COF ()'].mean()
                # max_mean_range = 1.1
                # if mean >= 0:
                #     max_cut_out = mean + max_mean_range # maksymalna wartość średniej ucinanej
                # else:
                #     max_cut_out = abs(mean) + max_mean_range # maksymalna wartość średniej ucinanej
                # mean_value = df.loc[(df['DAQ.COF ()'] <= max_cut_out) & (df['DAQ.COF ()'] >= -max_cut_out), 'DAQ.COF ()'].mean()
                # # Zastąpienie pików średnią
                # df['DAQ.COF ()'] = df['DAQ.COF ()'].apply(lambda x: mean_value if x > max_cut_out or x < -max_cut_out else x)

                # # Znajdź najniższą wartość w kolumnie "DAQ.COF ()"
                # min_value = df['DAQ.COF ()'].min()
                # # Jeśli najniższa wartość jest mniejsza od zera, zrób offset
                # if min_value < 0:
                #     df['DAQ.COF ()'] = df['DAQ.COF ()'] - min_value

                # Zmień nazwy kolumny
                df = df.rename(columns={'DAQ.COF ()': 'µ'})
                # Dodanie kolumn "Distance [m]" oraz "µ" i obliczenie tych danych dla Tribometru T11
                df = Rtec_calculations(df, os.path.basename(file_path))
                
                # Wybierz interesujące kolumny
                df = df[['Distance [m]', 'µ', 'Penetration Depth [µm]']]
            else:
                raise ValueError(f"\033[91m Brak funkcji programu typu [{mode}] dla tego tribometru: {tribometer_type} \033[0m")
        # Jak nie Obrotowy czy Posuwisto-zwrotny
        else:
            raise ValueError(f"\033[91m Brak wykrytego poprawnie rodzaju ruchu lub nie zdefiniowany, dla: {file_path} \033[0m")
        return df, tribometer_type, mode

    except Exception as e:
        print(f"\033[91m Błąd podczas przetwarzania pliku: {file_path}: {e} \033[0m")
        return None, None, None

# Funkcja oblicza najlepszą wartość ilości uśredniania próbek dla przebiegu µ i pd
#column_names = ['Distance [m]', 'µ', 'Penetration Depth [µm]']
def find_optimal_samples_average(data, column_names, min_sample=100, max_sample=500):
    min_sample_average = max(1, len(data) // max_sample)  # Wyznacz dolną granicę `sample_average` (MIN)
    # Dla max_sample = 500, jeśli len(data) = 1000, to min_sample_average = 2.
    max_sample_average = max(1, len(data) // min_sample)  # Wyznacz górną granicę `sample_average` (MAX)
    # Dla min_sample = 100, jeśli len(data) = 1000, to max_sample_average = 10.
    
    best_sample_average_µ = None
    best_sample_average_pd = None
    best_std_µ = float('inf')
    best_std_pd = float('inf')

    # Przechodzimy przez różne wartości `sample_average`
    for sample_average in range(min_sample_average, max_sample_average + 1):
        averaged_data_list = []

        # Tworzymy uśrednione dane
        for i in range(0, len(data), sample_average):
            subset = data.iloc[i:i+sample_average]
            
            if len(subset) < 2:
                continue
            
            µ_avg = subset['µ'].mean()
            pd_avg = subset['Penetration Depth [µm]'].mean()

            averaged_data_list.append({
                'Distance [m]': subset.iloc[0]['Distance [m]'],
                'µ': µ_avg,
                'Penetration Depth [µm]': pd_avg
            })
        
        averaged_data = pd.DataFrame(averaged_data_list, columns=column_names)
        std_µ = averaged_data['µ'].std()
        std_pd = averaged_data['Penetration Depth [µm]'].std()

        # Aktualizujemy najlepsze wartości `sample_average`
        if std_µ < best_std_µ:
            best_std_µ = std_µ
            best_sample_average_µ = sample_average
            
        if std_pd < best_std_pd:
            best_std_pd = std_pd
            best_sample_average_pd = sample_average
    
    return best_sample_average_µ, best_sample_average_pd  # Zwracamy krotkę (best_sample_average_µ, best_sample_average_pd)

# Funkcja dokonuje korekty wykresu zużycia liniowego aby zaczynał się od zera
# 1. spr. czy jest kolumna z danymi, 2. offset o najmniejszą wartość danych, 3. offset o najmniejsze lokalne minimum dla dodatnych danych, 
# 4. wykrycie piku ujemnego i go usunięcie z początka danych, 5. jak nie ma piku to offset o najniższą wartość w danych, 
# 6. Offset danych RAW aby przyrównać z danymi uśrednionymi, 7. usunięcie dużego peaku z początku danych uśrednionych, 
# 8. Inwersja dużego peaku z początku danych i offset w górę 9. wstaw 0 na początek danych, jak nie ma.
# df - dane uśrednione, data - dane RAW (czyste), percent - proc. danych do analizy piku, offset_raw - przyrównanie data do df, 
# erase_peak - usuwa znaczący pik danych, invert_peak - inwersja i dodanie offsetu piku danych
def process_penetration_depth(df, data, percent=0.1, offset_raw=1, erase_peak=0, invert_peak=1):
    if 'Penetration Depth [µm]' not in df.columns:  # Sprawdź, czy kolumna istnieje
        return df  # Zwróć df, jeśli kolumna nie istnieje
    else:
        # DLA WYKRESÓW UJEMNYCH
        # Oblicz i zastosuj offset dla najmniejszej wartości
        min_penetration_depth_1 = df['Penetration Depth [µm]'].min()  # Znajdź najmniejszą wartość
        if min_penetration_depth_1 < 0:  # Sprawdź, czy najmniejsza wartość jest mniejsza od 0
            df['Penetration Depth [µm]'] += abs(min_penetration_depth_1)  # Zastosuj offset do wszystkich wartości
            # ^^- przesunięcie wykresu w górę o najmniejszą wartość zbioru danych

        # DLA WYKRESÓW CAŁYCH DODATNICH (po powyższej korekcie)
        # Oblicz i zastosuj offset dla najmniejszego minimum lokalnego
        local_minima = (df['Penetration Depth [µm]'].shift(1) > df['Penetration Depth [µm]']) & (df['Penetration Depth [µm]'].shift(-1) > df['Penetration Depth [µm]'])
        min_penetration_depth_2 = df['Penetration Depth [µm]'][local_minima].min()  # Znajdź najmniejsze minimum lokalne
        if min_penetration_depth_2 > 0:  # Sprawdź, czy najmniejsze minimum lokalne jest większe od 0
            df['Penetration Depth [µm]'] += (-(min_penetration_depth_2))  # Zastosuj offset do wszystkich wartości
            # ^^- przesunięcie wykresu w dół o najmniejsze dodatnie minimum

        # DLA WYKRESÓW Z PIKIEM UJEMNYM NA POCZĄTKU (usuwanie ujemnych pików)
        # Sprawdzenie, czy pierwsza wartość w kolumnie Penetration Depth [µm] jest mniejsza od 0
        set_offset = False
        # tu wywalał błąd w niektórych przypadkach:
        # if df.loc[0, 'Penetration Depth [µm]'] < 0:
        #     # Oblicz procent wartości ze wszystkich wierszy (10% jako default dla zmiennej percent)
        #     limit = int(len(df) * percent)
        #     for index in range(1, min(limit, len(df))):
        #         if df.loc[index, 'Penetration Depth [µm]'] > 0:
        #             df.loc[:index-1, 'Penetration Depth [µm]'] = 0 # Ustaw 0 dla wszystkich wartości ujemnych przed tą wartością dodatnią
        #             break
        if df.iloc[0]['Penetration Depth [µm]'] < 0:
            # Oblicz procent wartości ze wszystkich wierszy (10% jako default dla zmiennej percent)
            limit = int(len(df) * percent)
            for index in range(1, min(limit, len(df))):
                if df.iloc[index]['Penetration Depth [µm]'] > 0:
                    df.iloc[:index, df.columns.get_loc('Penetration Depth [µm]')] = 0  # Ustaw 0 dla wszystkich wartości ujemnych przed tą wartością dodatnią
                    break
        # DLA WYKRESÓW UJEMNYCH ALE ROSNĄCYCH (cały czas albo prawie cały czas)
            else:
                # Jeśli wykres jest nadal ujemny powyżej limitu (10% jako default dla zmiennej percent)
                min_penetration_depth_3 = df['Penetration Depth [µm]'].min()  # Oblicz minimum z całej kolumny
                if min_penetration_depth_3 < 0:  # Sprawdź, czy minimum jest mniejsze od 0
                    df['Penetration Depth [µm]'] += abs(min_penetration_depth_3)  # Zastosuj offset, aby najmniejsza wartość była maksymalnie zero
                    set_offset = True
        # Dodatkowy offset wykresów rosnących (operacje już na wartościach dodatnich)
        # [oryginalne dane nie są uśredniane a początkowe wartości mogą się szybko zmieniać, więc poniższy kod to taki fix]
        if set_offset == True:
            min_penetration_depth_data_1 = data['Penetration Depth [µm]'].min()  # Oblicz minimum z całej kolumny ORYGINALNYCH danych
            if abs(min_penetration_depth_data_1) > abs(min_penetration_depth_1):
                difference = abs(min_penetration_depth_data_1) - abs(min_penetration_depth_1)
                df['Penetration Depth [µm]'] += abs(difference)  # Zastosuj offset, aby najmniejsza wartość była maksymalnie zero
                set_offset = False

        if erase_peak == 1: # (tylko dla wykresów powyżej lub równo z 0, czyli po korektach)
            # Usunięcie peaku zgodnie z zaleceniem Prof. MM
            min_value = df['Penetration Depth [µm]'].min()  # Znajdź najmniejszą wartość w kolumnie
            end_index = df[df['Penetration Depth [µm]'] == min_value].index  # Ustal indeks najmniejszej wartości
            if min_value <= 1.0 and not isinstance(end_index, pd.RangeIndex): # Tylko dla wartości mniejszych od 1 i nie dla RangeIndex
                if end_index[0] > 0: # Jeśli index nie jest zerowy (czyli nie pierwsza pozycja)
                    df.loc[:end_index[0], 'Penetration Depth [µm]'] = 0 # Ustaw wartości w kolumnie na 0 od 0 do pozycji end_index[0]
                elif len(end_index) > 1: # Jeśli będzie wiecej indeksów
                    df.loc[:end_index[-1], 'Penetration Depth [µm]'] = 0  # Ustaw wartości w kolumnie na 0 od 0 do pozycji ostatniego end_index

        elif erase_peak == 0 and invert_peak == 1: # (tylko dla wykresów powyżej lub równo z 0, czyli po korektach)
            # dodanie peaku jako inwersja i offset na początku danych
            consecutive_zero_min_index = 0  # Licznik wystąpień min_index = 0
            # Pętla while jest po to aby kod wykonać do puki nie wygładzi się przebiegów pseudo-sinusoidalnych
            while consecutive_zero_min_index < 3:  # Warunek zakończenia pętli - POTRÓJNY check, jeśli indeks się trzy razy nie zmieni to break
                min_value = df['Penetration Depth [µm]'].min()  # Znajdź najmniejszą wartość w kolumnie
                min_index = df['Penetration Depth [µm]'].idxmin()  # Znajdź indeks najmniejszej wartości
                end_index = df[df['Penetration Depth [µm]'] == min_value].index  # Ustal indeks najmniejszej wartości
                if min_value <= 0 and not isinstance(end_index, pd.RangeIndex):  # Tylko dla wartości mniejszych od 0 i nie dla RangeIndex
                    for i in range(len(end_index)):  # iterowanie po całym zakresie indeksów
                        df.loc[:end_index[i], 'Penetration Depth [µm]'] *= -1  # Inwertuj wartości w kolumnie od 0 do pozycji end_index[i]
                        min_value = df['Penetration Depth [µm]'].min()  # Znajdź najmniejszą wartość w kolumnie po inwersji
                        df['Penetration Depth [µm]'] += abs(min_value)  # Offsetuj całe dane, aby najmniejsza wartość była większa bądź równa 0
                if min_index == 0:  # Sprawdź, czy min_index jest równy 0
                    consecutive_zero_min_index += 1  # Zwiększ licznik
                else:
                    consecutive_zero_min_index = 0  # Zresetuj licznik, jeśli min_index nie jest równy 0

        if offset_raw == 1:
            # Offset danych RAW aby przyrównać z danymi uśrednionymi
            middle_index_df = len(df) // 2
            middle_index_data = len(data) // 2
            range_data = int(len(data) * 0.05)  # 5% zakresu danych
            # Oblicz indeksy dla uśrednionego odcinka danych
            start_index_data = max(0, middle_index_data - range_data)  # Indeks początkowy
            end_index_data = min(len(data), middle_index_data + range_data)  # Indeks końcowy
            # Oblicz średnią wartość dla uśrednionego odcinka danych
            average_penetration_depth = data['Penetration Depth [µm]'].iloc[start_index_data:end_index_data].mean()
            if df['Penetration Depth [µm]'].iloc[middle_index_df] != average_penetration_depth:
                difference = average_penetration_depth - df['Penetration Depth [µm]'].iloc[middle_index_df]
                data['Penetration Depth [µm]'] += -difference  # Dodaj offset danych w Penetration Depth [µm]

        # Upewnij się, że w zerowej linii "Penetration Depth [µm]" jest wartość 0
        if df.loc[0, 'Penetration Depth [µm]'] != 0:
            # Przesuń wartości w kolumnie 'Penetration Depth [µm]' w dół o 1
            df['Penetration Depth [µm]'] = df['Penetration Depth [µm]'].shift(1)
            # Dodaj nową wartość 0 na początek kolumny 'Penetration Depth [µm]'
            df.loc[0, 'Penetration Depth [µm]'] = 0

        # TODO kod nie usuwa dodatnich pików
        return df

# Funkcja uśrednia RAW Data według zmiennej best_sample_average i stasuje filtr Savitzky-Golay
def adjust_and_average_data(df, best_sample_average, window_u, window_pd):
    # Zainicjuj pustą listę na uśrednione dane
    averaged_data_list = []

    # Iteruj po zakresach danych co `sample_average` wierszy
    for i in range(0, len(df), best_sample_average):
        # Wybierz podzbiór danych dla danego zakresu
        subset = df.iloc[i:i+best_sample_average]

        # Oblicz średnią dla µ i Penetration Depth [µm] (dodatkowo zabezpieczenie 1: wartość µ nie może być ujemna)
        µ_avg = subset['µ'].mean()
        if µ_avg < 0:
            µ_avg = abs(µ_avg)
        pd_avg = subset['Penetration Depth [µm]'].mean()

        # Dodaj uśrednione wartości do listy
        averaged_data_list.append({
            'Distance [m]': subset.iloc[0]['Distance [m]'],  # Zachowaj oryginalną wartość Distance [m]
            'µ': µ_avg,
            'Penetration Depth [µm]': pd_avg
        })

    # Stwórz DataFrame z uśrednionymi danymi
    averaged_data = pd.DataFrame(averaged_data_list)

    # Druga filtracja
    averaged_data = Savitzky(averaged_data, window_u, window_pd)
    
    # Upewnij się, że w zerowej linii "µ" jest wartość 0
    if averaged_data.loc[0, 'µ'] != 0:
        # Przesuń wartości w kolumnie 'µ' w dół o 1
        averaged_data['µ'] = averaged_data['µ'].shift(1)
        # Dodaj nową wartość 0 na początek kolumny 'µ'
        averaged_data.loc[0, 'µ'] = 0

    # Upewnij się, że w zerowej linii "Distance [m]" jest wartość 0
    if averaged_data.loc[0, 'Distance [m]'] != 0:
        # Przesuń wartości w kolumnie 'Distance [m]' w dół o 1
        averaged_data['Distance [m]'] = averaged_data['Distance [m]'].shift(1)
        # Dodaj nową wartość 0 na początek kolumny 'Distance [m]'
        averaged_data.loc[0, 'Distance [m]'] = 0

    "Poniżej jest usuwanie powtórzeń wierszy, usuwanie o małej gęstości pomiarów, zostawienie kilku w przypadku braku danych"
    # Oblicz różnice bezwzględne między kolejnymi wierszami w 'Distance [m]'
    differences = averaged_data['Distance [m]'].diff().abs()
    if differences.any():
        # Usuń wiersze z danych względem kolumny "Distance [m]" tak, aby były maksymalnie z gęstością wartości co max 0.15
        filtered_data = averaged_data[(differences > 0.15) | (differences.isna())]
        # Sprawdź, czy po filtracji jest mniej niż dwa wiersze
        if filtered_data.shape[0] < 2:
            # Zachowaj tylko kilka wierszy z "Distance [m]" < 0.1
            averaged_data = averaged_data[averaged_data['Distance [m]'] < 0.1]
        else:
            # Jeśli jest więcej jak dwa wiersze to przypisz przefiltrowane dane do averaged_data
            averaged_data = filtered_data
    else:
        # Zostaw tylko dwa wiersze dla tych samych wartości Distance [m]
        averaged_data = averaged_data.groupby('Distance [m]').head(1 + 2)

    return averaged_data

# Funkcja dodaje ostatni pomiar z aproksymacji pomiarów, jeżeli brakuje ostatniego wiersza / linii
def approximate_last_measurement(averaged_data, original_data):
    if averaged_data.shape[0] < 2:  # Sprawdza liczbę wierszy
        print("\033[91mBrak wystarczających danych do przybliżenia\033[0m")  # Not enough data to approximate
        return averaged_data
    
    # TODO Remove last zero or negative values from rows in averaged_data (OPTIONAL)

    # Check if last two Distance [m] values are the same
    if averaged_data['Distance [m]'].iloc[-1] == averaged_data['Distance [m]'].iloc[-2]:
        return averaged_data  # Skip if last two Distance [m] values are the same

    # Get the original last Distance [m] value
    original_last_distance = original_data['Distance [m]'].iloc[-1]

    # Find the index of original_data where Distance [m] is the original_last_distance
    original_last_index = original_data.index[original_data['Distance [m]'] == original_last_distance]

    # Check if the index is not empty
    if not original_last_index.empty:
        # Get the index location
        original_last_index = original_last_index.values[0]

        # Create a new row with the last averaged values and original last Distance [m]
        new_row = {
            'Distance [m]': original_data.loc[original_last_index, 'Distance [m]'],
            'µ': averaged_data['µ'].iloc[-1],
            'Penetration Depth [µm]': averaged_data['Penetration Depth [µm]'].iloc[-1] if 'Penetration Depth [µm]' in averaged_data else None
        }

        # Concatenate the new row to averaged_data if Distance [m] value is not duplicated
        if new_row['Distance [m]'] not in averaged_data['Distance [m]'].values:
            averaged_data = pd.concat([averaged_data, pd.DataFrame([new_row])], ignore_index=True)

    return averaged_data

# Funkcja generująca wykresy z danych
def generate_combined_xlsx(csv_files, output_xlsx, title_from_text):
    # Utwórz nowy Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Data"
    text_in_brackets = ""
    cleaned_filename = ""
    offset_row = 0

    col_offset = 0
    row_offset = 0

    for csv_file in csv_files:
        df = pd.read_csv(csv_file)
            
        filename = os.path.basename(csv_file)

        # Tekst nad wszystkimi danymi
        header_text = f"Poniżej znajdują się dane z tribometru sformatowane i uśrednione dynamicznie"

        # Zapisz tekst nad wszystkimi danymi
        ws.cell(row=1, column=1, value=header_text)

        # Znajdź tekst w nawiasie w zmiennej filename aby ustawić nazwę serii danych dla wykresów
        match = re.search(r'\((.*?)\)', filename)
        if match:
            text_in_brackets = match.group(1)
            cleaned_filename = filename.replace(text_in_brackets, "").replace('(', "").replace(')', "").replace('.csv', "")
            ws.cell(row=2, column=col_offset + 3, value=text_in_brackets)
        else:
            cleaned_filename = filename.replace('.csv', "").replace('(', "").replace(')', "")
            ws.cell(row=2, column=col_offset + 3, value=filename)

        # Zapisz nazwę pliku nad nagłówkami danych
        ws.cell(row=2, column=col_offset + 1, value=cleaned_filename)

        # Zapisz nagłówki danych
        for col_idx, col_name in enumerate(df.columns):
            ws.cell(row=3, column=col_idx + 1 + col_offset, value=col_name)

        # Zapisz dane z DataFrame do arkusza, z uwzględnieniem przesunięcia kolumn
        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx + 4, column=c_idx + 1 + col_offset, value=value)

        # Pobierz tytuł osi X z trzeciego wiersza
        x_axis_title = ws.cell(row=3, column=col_offset + 1).value

        # Ustal maksymalną wartość dla osi X
        max_distance = df['Distance [m]'].max()
        x_axis_max = int(round(max_distance))

        # Dodaj offset na dane (fix glitch)
        if title_from_text == 1:
            offset_row = -1

        # Tworzenie wykresu dla µ
        chart = ScatterChart() # ostra linia
        #chart = LineChart() # wygładzona linia
        chart.title = f"{cleaned_filename + ' ' + text_in_brackets} - µ"
        chart.style = 2  # Nowy rozmiar czcionki
        chart.x_axis.title = x_axis_title
        if x_axis_max < 10:
            chart.x_axis.scaling.max = 100
        else:
            chart.x_axis.scaling.max = x_axis_max
        chart.x_axis.majorUnit = int(round(x_axis_max / 5.0))
        x_data = Reference(ws, min_col=col_offset + 1, min_row=4, max_row=len(df) + 3)
        y_data = Reference(ws, min_col=col_offset + 2, min_row=4 + offset_row, max_row=len(df) + 3)
        series = Series(y_data, x_data, title_from_data=title_from_text, title=text_in_brackets)
        chart.series.append(series)
        # Pobierz tytuł osi Y dla wykresu dla µ (kolumna po prawej stronie od Distance [m])
        # y_axis_title_mu = ws.cell(row=3, column=col_offset + 2).value
        chart.y_axis.title = "Friction coefficient" # y_axis_title_mu
        ws.add_chart(chart, f"B{25 + 6 + row_offset}")

        if 'Penetration Depth [µm]' in df.columns:
            # Tworzenie wykresu dla Penetration Depth [µm]
            chart_pd = ScatterChart()
            #chart_pd = LineChart()
            chart_pd.title = f"{cleaned_filename + ' ' + text_in_brackets} - Penetration Depth [µm]"
            chart_pd.style = 2  # Nowy rozmiar czcionki
            chart_pd.x_axis.title = x_axis_title
            if x_axis_max < 10:
                chart_pd.x_axis.scaling.max = 100
            else:
                chart_pd.x_axis.scaling.max = x_axis_max
            chart_pd.x_axis.majorUnit = int(round(x_axis_max / 5.0))
            x_data_pd = Reference(ws, min_col=col_offset + 1, min_row=4, max_row=len(df) + 3)
            y_data_pd = Reference(ws, min_col=col_offset + 3, min_row=4 + offset_row, max_row=len(df) + 3)
            series_pd = Series(y_data_pd, x_data_pd, title_from_data=title_from_text, title=text_in_brackets)
            chart_pd.series.append(series_pd)
            # Pobierz tytuł osi Y dla wykresu dla Penetration Depth [µm] (kolumna po prawej stronie od µ)
            # y_axis_title_pd = ws.cell(row=3, column=col_offset + 3).value
            chart_pd.y_axis.title = "Linear wear [µm]" # y_axis_title_pd
            ws.add_chart(chart_pd, f"M{25 + 6 + row_offset}")

        # Dodaj pustą kolumnę jako separator
        col_offset += len(df.columns) + 1
        
        # Przesuń wskaźnik wiersza na pozycję pod wykresami
        row_offset += 15

    # Zapisz plik xlsx
    wb.save(output_xlsx)
    print(f"Dane zostały zapisane do pliku {output_xlsx}")

def main():
    print("Program do obróbki danych z tribometru: TRB3, Nano TRB, T11, Rtec")
    print("Wczytuje pliki .txt i .csv z folderu z programem")
    print("Pliki te powinny mieć w nazwie nawias () a w nim nazwę serii")
    print("Pliki z Rtec powinny mieć w nazwie prędkość liniową np. 0.1m-s a dla T11 dodatkowo obciążenie np. 10N\n")
    folder_path = '.'
    csv_files = []
    csv_files_raw = []
    config_path = "config.ini"
    if not os.path.isfile(config_path):
        config_path = "_config.ini"
    if os.path.isfile(config_path):
        config = load_config(config_path) # Wczytaj dane z configu
        offset_raw = config['offset_raw']
        title_from_text = config['title_from_text']
        erase_peak = config['erase_peak']
        invert_peak = config['invert_peak']
        default_window_length_u = config['default_window_length_u']
        default_window_length_pd = config['default_window_length_pd']
        min_sample = config['min_sample']
        max_sample = config['max_sample']
    else:
        min_sample, max_sample, default_window_length_u, default_window_length_pd, title_from_text, offset_raw, erase_peak, invert_peak = ask_user_for_variables() # Wczytaj dane od użytkownika
    for filename in os.listdir(folder_path):
        if filename.endswith(".txt") or filename.endswith(".csv"):
            file_path = os.path.join(folder_path, filename)
            data, tribometer_type, mode = read_and_process_file(file_path) # Wczytaj dane z plików
            if data is not None:
                # Przetwórz dane przy użyciu najlepszej wartości sample_average (do wyboru µ lub pd)
                best_sample_average_µ, best_sample_average_pd = find_optimal_samples_average(data, column_names=['Distance [m]', 'µ', 'Penetration Depth [µm]'], min_sample=min_sample, max_sample=max_sample)
                averaged_data = adjust_and_average_data(data, best_sample_average_µ, default_window_length_u, default_window_length_pd)
                
                # Korekta zużycia liniowego
                averaged_data = process_penetration_depth(averaged_data, data, 0.1, offset_raw, erase_peak, invert_peak) # df=averaged_data, data=data, percent=0.1, offset_raw=0, erase_peak=0, invert_peak=1

                # Approximate the last value
                approximated_data = approximate_last_measurement(averaged_data, data)

                # Nazwa plików CSV odpowiadających nazwie pliku tekstowego oryginalnego + spacja
                output_file = os.path.splitext(filename)[0] + ' .csv'
                output_file_raw = "raw_" + output_file # to samo co wyżej, ale z przedrostkiem "raw_"
                
                if "Nano" in tribometer_type: # "Nano" - usuń kolumnę pd dla nano tribometru
                    approximated_data = approximated_data.drop(columns=['Penetration Depth [µm]']) # z obrobionych
                    data = data.drop(columns=['Penetration Depth [µm]']) # z raw

                # Zapisz wynik do pliku CSV
                try:
                    approximated_data.to_csv(output_file, index=False, float_format='%.3f') # finalne dane wyjściowe
                    data.to_csv(output_file_raw, index=False, float_format='%.3f') # dane tylko wstępnie obrobione
                    csv_files.append(output_file)
                    csv_files_raw.append(output_file_raw)
                    print(f"[{tribometer_type}][{mode}] średnia dla µ: {best_sample_average_µ}, plik: {filename}")
                except Exception as e:
                    print(f"\033[91mBłąd podczas zapisywania pliku {output_file}: {e} \033[0m")
                
    # Generowanie pliku xlsx ze wszystkimi danymi
    status = 0
    if csv_files:
        try:
            generate_combined_xlsx(csv_files, "combined_data.xlsx", title_from_text)
            status = 1
        except Exception as e:
            status = 0
            print(f"\033[91mBłąd podczas zapisywania pliku combined_data.xlsx: {e} \033[0m")
        for file in csv_files:
            if os.path.exists(file):
                os.remove(file)
        # Sprawdź, czy wszystkie pliki zostały usunięte
        all_removed = True
        for file in csv_files:
            if os.path.exists(file):
                all_removed = False
                print(f"\033[38;5;214m Plik {file} nadal istnieje. \033[0m")
        if all_removed:
            print("Wszystkie pliki zostały usunięte.")
        # Generowanie pliku xlsx ze wszystkimi danymi
    if csv_files_raw:
        try:
            generate_combined_xlsx(csv_files_raw, "combined_data_raw.xlsx", title_from_text)
            status = 1
        except Exception as e:
            status = 0
            print(f"\033[91mBłąd podczas zapisywania pliku combined_data_raw.xlsx: {e} \033[0m")
        for file in csv_files_raw:
            if os.path.exists(file):
                os.remove(file)
        # Sprawdź, czy wszystkie pliki zostały usunięte
        all_removed = True
        for file in csv_files_raw:
            if os.path.exists(file):
                all_removed = False
                print(f"\033[38;5;214mPlik {file} nadal istnieje. \033[0m")
        if all_removed:
            print("Wszystkie pliki raw zostały usunięte.")
    if status == 1: print("\033[92mWszystkie dane z WCZYTANYCH plików zostały zapisane poprawnie \033[0m")
    else: print("\033[91mNie wszystkie dane zostały zapisane poprawnie z powodu powyższych błędów \033[0m")
    countdown = 5
    while countdown > 0:
        print(f"Okno konsoli zostanie zamknięte za {countdown} s, naciśnij ESC by anulować zamknięcie", end="\r")  # Wyświetl odliczanie w tej samej linii
        if msvcrt.kbhit() and msvcrt.getch() == b'\x1b':  # Sprawdź, czy naciśnięto klawisz Escape
            print("\n Anulowano zamknięcie, teraz naciśnij dowolny klawisz, aby zamknąć konsolę.")
            msvcrt.getch()  # Czeka na naciśnięcie dowolnego klawisza
        time.sleep(1)
        countdown -= 1

if __name__ == "__main__":
    main()

# TODO dodać komunikat że nie ma plików do wczytania (pusty folder)
# TODO poprawić komunikat braku zmiennych w nazwie pliku (usunąć nawias kwadratowy i standardowe zmienne 0.1m/s i 10N)
# TODO dodać procent przetwarzania danych
